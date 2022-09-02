# -*- coding: utf-8 -*-

import pyodbc as p
import openpyxl, pprint
import os
import sys
import datetime
from openpyxl.styles import NamedStyle
from django.http import HttpResponse
import csv
import ftplib
import shutil
from . import stringConnexao

        
def def002_cadastrarPastas(planilha,operacao,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames
                
        sheet = wb.get_sheet_by_name(sheets[0])

        sql_command1 =   """
        INSERT INTO dbo.tab_incluirpastas (id_operador,lote,
        ordem,
        num_processo,
        escritorio,
        cod_cliente,
        dat_abertura,
        estado,
        comarca,
        supervisor,
        responsavel,
        estagiario,
        conveniado,
        publicando,
        exadverso,
        autor,
        vitima,
        reu,
        dat_citacao,
        dat_distribuicao,
        orgao,
        vara,
        rito,
        valor
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """

        qtde_itens=0
        row=2
        erro=0
        while row<sheet.max_row+1:
            qtde_itens+=1
            ccA = sheet['A' + str(row)].value
            ccB = sheet['B' + str(row)].value
            ccC = sheet['C' + str(row)].value
            ccD = sheet['D' + str(row)].value
            ccE = sheet['E' + str(row)].value
            ccF = sheet['F' + str(row)].value
            ccG = sheet['G' + str(row)].value
            ccH = sheet['H' + str(row)].value
            ccI = sheet['I' + str(row)].value
            ccJ = sheet['J' + str(row)].value
            ccK = sheet['K' + str(row)].value
            ccL = sheet['L' + str(row)].value
            ccM = sheet['M' + str(row)].value
            ccN = sheet['N' + str(row)].value
            ccO = sheet['O' + str(row)].value
            ccP = sheet['P' + str(row)].value
            ccQ = sheet['Q' + str(row)].value
            ccR = sheet['R' + str(row)].value
            ccS = sheet['S' + str(row)].value
            ccT = sheet['T' + str(row)].value
            ccU = sheet['U' + str(row)].value
            ccV = sheet['V' + str(row)].value
            try:
                if (ccA):
                    db_cursor.execute(sql_command1, idop,lote, 
                    ccA, 
                    ccB, 
                    ccC,
                    ccD,
                    ccE,
                    ccF, 
                    ccG,
                    ccH,
                    ccI,
                    ccJ,
                    ccK,
                    ccL,
                    ccM,
                    ccN,
                    ccO,
                    ccP,
                    ccQ,
                    ccR,
                    ccS,
                    ccT,
                    ccU,
                    ccV
                    )
                    db_connection.commit()

            except p.IntegrityError:
                print("Erro na inclusao.")
            row+=1

        sqlExecSP="""\
        Exec dbo.s036_cadastramentoDePastas @id_operador=?, @lote=?, @opcao=?
        """
        params = (idop,lote,'X')
        db_cursor.execute(sqlExecSP, params)
        
        db_cursor.close()
        del db_cursor
        db_connection.close()
        gerar_arquivo_erro(lote)
        
        

def gerar_arquivo_erro(lote):

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()
        sql_command =   """
        SELECT e.id_seq,e.lote,e.linha,e.erro,e.cod_cliente, e.mensagem
        FROM errosImportacao e,loteImportacao l 
        WHERE e.lote=l.lote  ORDER BY e.id_seq
        """
        try:
            ErrosImportacao.objects.all().delete()
            db_cursor.execute(sql_command)
            row = db_cursor.fetchone() 

            while row:
                erro = ErrosImportacao(lote=row[1],linha=row[2],erro=row[3],cod_cliente=row[4], mensagem=row[5])
                erro.save()
                row = db_cursor.fetchone() 
        except p.IntegrityError:
            print ("Erro na inclusao")
            
        db_cursor.close()
        del db_cursor
        db_connection.close()



def def005_sinistro(planilha,operacao,id_user):
    lote = str(datetime.datetime.now().today())[0:19]
    idop = id_user
    print('planilha: ',planilha)

    string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()

    wb = openpyxl.load_workbook(planilha)
            
    sheet = wb.get_sheet_by_name('Plan1')
    sql_command =   """
    INSERT INTO dbo.tab_sinistro (
    id_operador,
    lote,
    cod_cliente,
    sinistro,
    cod_status,
    vitima,
    cpf,
    datnasc,
    cobertura,
    datsinistro,
    categoria,
    uf,
    municipio,
    renavam,
    placa,
    chassi,
    placa_mercosul
    )
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    row=2
    erro=0
    qtde_itens=0
    #sheet.get_highest_row()
    while row<sheet.max_row+1:
        qtde_itens+=1
        colA = sheet['A' + str(row)].value #
        colB = sheet['B' + str(row)].value #
        colC = sheet['C' + str(row)].value #
        colD = sheet['D' + str(row)].value #
        colE = sheet['E' + str(row)].value #
        colF = sheet['F' + str(row)].value
        colG = sheet['G' + str(row)].value #
        colH = sheet['H' + str(row)].value
        colI = sheet['I' + str(row)].value #
        colJ = sheet['J' + str(row)].value #
        colK = sheet['K' + str(row)].value #
        colL = sheet['L' + str(row)].value #
        colM = sheet['M' + str(row)].value #
        colN = sheet['N' + str(row)].value #
        colO = sheet['O' + str(row)].value #
        try:
            colI = colI[:3]
            db_cursor.execute(sql_command, 
            idop,
            lote, 
            colA, 
            colB, 
            colC, 
            colD, 
            colE,
            colF,
            colG,
            colH,
            colI,
            colJ,
            colK,
            colL,
            colM,
            colN,
            colO)
            
            db_connection.commit()
        except p.IntegrityError:
            print("Erro na inclusao.")
        row+=1

    sqlExecSP="""\
    Exec dbo.s044_tab_sinistro @id_operador=?, @lote=?, @operacao=?, @programa=?
    """
    params = (idop,lote,'X','def005')
    db_cursor.execute(sqlExecSP, params)
            
    db_cursor.close()
    del db_cursor
    db_connection.close()       
        


def hon_conv():
    conveniados =  Conveniado.objects.all()
    for conv in conveniados:
        conv.c209001=0
        conv.c209002=0
        conv.save()

    string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()
    sql_command =   """
        select a.id_advogado,
        (select count(*) from hon_solicitacao h1
         where h1.id_advogado=a.id_advogado
            and h1.mes_base='209001') as mes209001,
        (select count(*) from hon_solicitacao h2
         where h2.id_advogado=a.id_advogado
            and h2.mes_base='209002') as mes209002
                from advogados a
                 where 
                    a.id_advogado in (15786,6757,42,14170,13059, 9344,14449,15055,2330,1982, 107)
        """
    try:
        db_cursor.execute(sql_command)
        row = db_cursor.fetchone() 
        while row:
            id = int(row[0])
            qt = int(row[1])
            #print (int(row[0]),int(row[1]))
            conveniado = Conveniado.objects.get(cid=id)
            conveniado.c209001=row[1]
            conveniado.c209002=row[2]
            conveniado.save()
            row = db_cursor.fetchone() 
    except p.IntegrityError:
        print ("Erro na inclusao")
            
    db_cursor.close()
    del db_cursor
    db_connection.close()

def alterconv(request, opcao, id_conv):     
    current_user = request.user.iduser
    conveniado=Conveniado.objects.get(pk=id_conv)
    id_conveniado=conveniado.cid
    tabela='X'

    string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()   string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()


    if opcao=="1":
        sqlExecSP="""\
        Exec dbo.TrocarFase_209001_209002 @idadv=?, @id_operador=?, @tabela=?
        """
        params = (id_conveniado, current_user, tabela)
        db_cursor.execute(sqlExecSP, params)


    else:
        sqlExecSP="""\
        Exec dbo.TrocarFase_209002_mes_corrente @idadv=?, @id_operador=?, @tabela=? 
        """ 
        params = (id_conveniado, current_user, tabela)
        db_cursor.execute(sqlExecSP, params)
            
    db_cursor.close()
    del db_cursor
    db_connection.close()

    
    
def incluirDocTabelaBanco(data_inicial,data_final,id_documento, usuario, lote):
    string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()
    sqlExecSP="""\
    Exec dbo.preparaDownload @data1=?, @data2=?, @id_doc=?, @id_user=?, @lote=?
    """
    
    params = (data_inicial,data_final, id_documento, usuario, lote)
    db_cursor.execute(sqlExecSP, params)
    
    db_cursor.close()
    del db_cursor
    db_connection.close()



