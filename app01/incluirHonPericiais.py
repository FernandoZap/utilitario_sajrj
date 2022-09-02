# -*- coding: utf-8 -*-

import pyodbc as p
import openpyxl
import os
import sys
import datetime
from . import stringConnexao



def incluirHonPericiais(planilha,operacao,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()


        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames
        sheet0 = sheets[0]
        sheet = wb.get_sheet_by_name(sheet0)
        
        sql_command1 =   """
        INSERT INTO dbo.tab_honorariospericiais (id_operador,lote,operacao,chave,data_evento,data_pagamento,valor,id_status)        
         VALUES (?,?,?,?,?,?,?,?)
        """

        sql_command2 =   """
        INSERT INTO dbo.tab_honorariospericiais (id_operador,lote,operacao,chave,data_evento,condicao,valor,id_status)        
         VALUES (?,?,?,?,?,?,?,?)
        """

        row=2
        erro=0
        ccA=operacao
        qtde_itens=0
            

        row=2
        ccA=operacao
        erro=0
        while row<sheet.max_row+1 and ccA==operacao and erro==0:
            qtde_itens+=1
            ccA = sheet['A' + str(row)].value # DC 
            ccB = sheet['B' + str(row)].value # Código da pasta
            ccC = sheet['C' + str(row)].value # Id do tipo da decisao
            ccD = sheet['D' + str(row)].value # Id da decisao
            ccE = sheet['E' + str(row)].value # Id da decisao
            ccF = sheet['F' + str(row)].value # Id da decisao
            ccG = sheet['G' + str(row)].value # Id da decisao

            ccA=ccA.strip()
            try:
                if (ccA==operacao):
                    if(ccD):
                        db_cursor.execute(sql_command1, idop,lote, ccA, ccB, str(ccC)[0:10], str(ccD)[0:10], ccF, ccG)
                        db_connection.commit()
                    else:
                        db_cursor.execute(sql_command2, idop,lote, ccA, ccB, str(ccC)[0:10],  ccE, ccF, ccG)
                        db_connection.commit()

            except p.IntegrityError:
                print("Erro na inclusao.")
            row+=1
            ccA = sheet['A' + str(row)].value # DC 
            if ccA:
                ccA = ccA.strip()

        sqlExecSP="""\
        Exec dbo.s018_cadastroHonPericiais @id_operador=?, @lote=?, @flag=?
        """
        flag='X'
        params = (idop,lote,flag)
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)
        
        db_cursor.close()
        del db_cursor
        db_connection.close()


