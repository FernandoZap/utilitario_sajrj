import pyodbc as p
import openpyxl
import datetime
import os
from . import stringConnexao


def ler_plan_pastas(planilha, lote, id_operador, programa):

    string_db=stringConnexao.strSqlServer()
    db_connection = p.connect(string_db)
    db_cursor = db_connection.cursor()


    wb = openpyxl.load_workbook(planilha)
    sheets = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheets[0])

    row = 2
    chave='codigo_cliente'
    if len(str(sheet['A' + str(row)].value))==10:
        if str(sheet['A' + str(row)].value)[4:5]=='-':
            chave='codigo_saj'

    fim=True
    row = 2

    if (programa=="view003"):
        sql_command =   """
        INSERT INTO dbo.tab_amarracao_01 (chave,id_operador,lote,programa)        
        VALUES (?,?,?,?)
        """
    elif (programa=="view001"):
        sql_command =   """
        INSERT INTO dbo.tab_dadosGerais (chave,id_operador,lote,programa)        
        VALUES (?,?,?,?)
        """
    elif (programa=="view011"):
        sql_command =   """
        INSERT INTO dbo.tab_consultas (chave,id_operador,lote)        
        VALUES (?,?,?)
        """

    finalizar=0



    while row<sheet.max_row+1 and finalizar==0:
        if (chave=='codigo_cliente'):
            item = str(sheet['A' + str(row)].value)
        elif (chave=='codigo_saj'):
            item = sheet['A' + str(row)].value
        if (programa=="view011"):
            db_cursor.execute(sql_command, item, id_operador,lote)
        else:
            db_cursor.execute(sql_command, item, id_operador,lote,programa)
        db_connection.commit()
        row+=1
        if sheet['A' + str(row)].value==None:
            finalizar=1
        elif sheet['A' + str(row)].value=='':
            finalizar=1

    if (programa=="view003"):
        sqlExecSP="""\
        Exec dbo.proc_tab_amarracao_01 @id_operador=?, @lote=?, @operacao=?
        """
    elif (programa=="view001"):
        sqlExecSP="""\
        Exec dbo.proc_tab_dadosGerais @id_operador=?, @lote=?, @operacao=?
        """
    elif (programa=="view011"):
        sqlExecSP="""\
        Exec dbo.py009_tab_honorarios @id_operador=?, @lote=?, @operacao=?
        """


    params = (id_operador,lote,'xxxxx')
    db_cursor.execute(sqlExecSP, params)

    db_cursor.close()
    del db_cursor
    db_connection.close()   
