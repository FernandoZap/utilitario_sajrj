import pyodbc as p
import openpyxl
import datetime
import mysql.connector
from . import stringConnexao
import os


def ler_plan_pastas(planilha, lote, id_operador, programa):
    wb = openpyxl.load_workbook(planilha)
    sheets = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheets[0])

    row = 2
    chave = 'codigo_cliente'
    if len(str(sheet['A' + str(row)].value)) == 10:
        if str(sheet['A' + str(row)].value)[4:5] == '-':
            chave = 'codigo_saj'

    dict_strings=stringConnexao.strMySql()
    cnx1 = mysql.connector.connect(user=dict_strings.get('user'), password=dict_strings.get('password'),
                                   host=dict_strings.get('host'),
                                   database=dict_strings.get('database'))
    cursor = cnx1.cursor()

    fim = True
    row = 2

    """

    if (programa=="view003"):
        add_tab_tabela = ("INSERT INTO tab_amarracao_01 "
               "(chave, id_operador, lote, programa) "
               "VALUES (%s, %s, %s, %s)")
    elif (programa=="dadosGerais"):
        add_tab_tabela = ("INSERT INTO tab_dadosGerais "
               "(chave, id_operador, lote, programa) "
               "VALUES (%s, %s, %s, %s)")
    """
    add_tab_tabela = (
        "INSERT INTO tab_dadosGerais (chave, id_operador, lote, programa) VALUES (%s, %s, %s, %s)")

    while row < sheet.max_row+1:
        if (chave == 'codigo_cliente'):
            item = str(sheet['A' + str(row)].value)
        elif (chave == 'codigo_saj'):
            item = sheet['A' + str(row)].value
        data_campos = (item, id_operador, lote, programa)
        cursor.execute(add_tab_tabela, data_campos)
        row += 1
    cnx1.commit()

    cursor.close()
    cnx1.close()

