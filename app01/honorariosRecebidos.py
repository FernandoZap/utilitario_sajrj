# -*- coding: utf-8 -*-

import pyodbc as p
import openpyxl
import os
import sys
import datetime
from . import stringConnexao

def incluir(planilha,operacao,fase,aba_planilha,current_user):
        lote = str(datetime.datetime.now().today())[0:19]

        idop = current_user
        fase = int(fase)

        string_db=stringConnexao.strSqlServer()
        db_connection = p.connect(string_db)
        db_cursor = db_connection.cursor()

        wb = openpyxl.load_workbook(planilha)
        sheets = wb.sheetnames

        sheet0 = sheets[int(aba_planilha)]
        
        sheet = wb.get_sheet_by_name(sheet0)


        sql_command =   """
        INSERT INTO dbo.tab_sitCobranca (id_operador,lote,cod_cliente,fase,planilha)        
         VALUES (?,?,?,?,?)
        """

        erro=0
        qtde_itens=0
        retorno = True
        row=1
        erro=0

        while row<sheet.max_row+1 and erro==0 and row<2000:
            qtde_itens+=1
            ccA = str(sheet['A' + str(row)].value) # codigo do cliente
            ccC = str(sheet['C' + str(row)].value) # fase

            ccA = ccA.strip()
            ccC = ccC.strip()
            ccA = ccA.upper()

            if fase==4 and ccC[0]=='3':
                ccC='3ª Fase/Exito'


            try:
                db_cursor.execute(sql_command, idop,lote, ccA, ccC, fase)
                db_connection.commit()
            except p.IntegrityError:
                print("Erro na inclusao.")
            ccA = sheet['A' + str(row)].value # DC 
            row+=1
            if ccA:
                ccA = str(ccA)
                ccA = ccA.strip()
                ccA = ccA.upper()
            else:
                erro=1

        sqlExecSP="""\
        Exec dbo.py005_inserirHonorarios @id_operador=?, @lote=?, @flag=?
        """

        params = (idop,lote,'INSERIR')
        db_connection.autocommit=True
        db_cursor.execute(sqlExecSP, params)
        
        db_cursor.close()
        del db_cursor
        db_connection.close()
